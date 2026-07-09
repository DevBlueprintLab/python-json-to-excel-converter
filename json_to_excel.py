# Json-Excel converter
import json, openpyxl, sys
from openpyxl.styles import Font, Alignment, PatternFill, Side, Border
from openpyxl.utils import get_column_letter
from pathlib import Path
print("=" * 30)
print("JSON to Excel Converter")
print("=" * 30)
# ask for valid json path
while True:
    json_file_path = Path(input('Enter JSON file path: ').strip())
    if json_file_path.is_file():
        break
    else:
        print('File not found. try again.')
# make new folder for converted excel
excel_folder = json_file_path.parent /'converted'
excel_folder.mkdir(parents = True, exist_ok = True )
excel_file_path = excel_folder / f'{json_file_path.stem}.xlsx'
with open(json_file_path, encoding = 'utf-8') as file:
    try:
        json_data = json.load(file)
    except json.JSONDecodeError :
        print('Invalid JSON file! Program exiting...')
        sys.exit()
#check if Json file structure
if not isinstance(json_data, list):
    print('Json file is not a list. Exiting...')
    sys.exit()
if not json_data:
    print('File is empty. Exiting...')
    sys.exit()
for item in json_data:
    if not isinstance(item, dict):
        print('Items are not dictionaries. Exiting...')
        sys.exit()
headers = set(json_data[0].keys())
for item in json_data:
    if set(item.keys()) != headers:
        print('Dictionaries don\'t have the same keys!\nExiting...')
        sys.exit()
# create workbook
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = 'Converted data'
# writing converting json file to excel
print('Converting to excel file...')
for col, key in enumerate(headers, start = 1):
    sheet.cell(row = 1, column = col ).value = key
for row, item in enumerate(json_data, start = 2):
    for col, key in enumerate(headers, start = 1):
        sheet.cell(row = row, column = col).value = item[key]
# format Excel file
header_font = Font( size=12, color = '00FFFFFF', bold = True)
header_fill = PatternFill(fill_type = 'solid', end_color = '00003366')
value_alignment = Alignment(horizontal = 'center', vertical = 'center')
bold_font = Font(bold = True)
right_border_side = Side(border_style = 'thick', color = '00000000')
right_border = Border(right = right_border_side)
sheet.auto_filter.ref = sheet.dimensions
sheet.freeze_panes = 'A2'
# apply boder, alignment and header styles
for col in range(1, sheet.max_column +1):
    sheet.cell(row = 1, column = col).font = header_font
    sheet.cell(row = 1, column = col).fill = header_fill
    sheet.cell(row = 1, column = col).alignment = value_alignment
for col_num in range(1, sheet.max_column +1):
    for row_num in range(2, sheet.max_row +1):
        sheet.cell(row= row_num, column = col_num).alignment = value_alignment
        sheet.cell(row =row_num, column = 1).border = right_border
        sheet.cell(row =row_num, column = 1).font = bold_font
# set automatic cell dimension
for col in range(1, sheet.max_column +1):
    max_length = 0
    for row in range(1, sheet.max_row + 1):
        value = sheet.cell(row = row, column = col).value
        if value:
            max_length = max(max_length, len(str(value)))
    letter = get_column_letter(col)
    sheet.column_dimensions[letter].width = (max_length + 5)
wb.save(excel_file_path)
print('\nConversion finished!\n')
print(f'Converted excel rows: {len(json_data)}')
print(f'Excel saved at {excel_file_path}\n')